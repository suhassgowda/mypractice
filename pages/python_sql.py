import json

import mysql.connector as connector
import openpyxl


class DbHelper:

    def __init__(self, driver):
        self.mydb = connector.connect(host="localhost", user="root", passwd="Test@123", database="testing",
                                      auth_plugin="mysql_native_password")
        self.driver = driver

    def retrieve_data_database(self):
        mycursor = self.mydb.cursor()

        mycursor.execute("select * from exports")
        records = mycursor.fetchall()
        record = []
        for row in records:
            record.append(row)
        for data in record:
            print(data, end=",")
            # print(data, end="[]")
            print("\b\b", end="")
            print("")

            return data

    def retrieve_data_excel(self):
        wb = openpyxl.load_workbook("D:\\file\\exports.xlsx")
        sh1 = wb['exports']
        row = sh1.max_row
        column = sh1.max_column
        data_excel = []

        for i in range(2, row+1):
            for j in range(1, column+1):
                data_excel.append(sh1.cell(i, j).value)
                # print(sh1.cell(i, j).value)
        for data_file in data_excel:
            print(data_file, end=",")
            # print(data_file, end="[]")
            print("\b", end="")
            print("")

            return data_file

    def compare_database_excel_data(self, data, data_file):
        data_db = data
        excel_data = data_file
        # print(data_db)
        # print(excel_data)
        if data_db.__contains__(excel_data):
            print("SQL data present in excel file")
        else:
            print("Not present")

    def read_business_rules_from_DMH_API_Response(self):
        """
                Read the business rules from the DMH API response
        """
        with open('DMH_API_Response.json', errors="ignore") as f:
            data = json.load(f)
        object = json.dumps(data)
        json_object = json.loads(object)
        business_rules = json_object['process']['pipeline_params']['transformation_config']['age_bands'][0]
        return business_rules

    def values_fetched_One(self, business_rules):
        print(business_rules)
        data_value = []
        for list in business_rules:
            # print(list)
            data_value.append(list)
            print(list)
        print(data_value)


